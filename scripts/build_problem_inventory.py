from __future__ import annotations

import csv
import os
import re
import shutil
import subprocess
from collections import defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
NOTE_DIR = ROOT / "note"
STUDY_DIR = ROOT / "study"
COVERAGE_DIR = STUDY_DIR / "coverage"

PDFTOTEXT = os.environ.get("PDFTOTEXT") or shutil.which("pdftotext")

CHAPTER_MAP = {
    "02-01 生命保険会計.md": "保険2第1章 生命保険会計",
    "02-03 契約者配当.md": "保険2第3章 契約者配当",
    "02-04 リスク管理.md": "横断 リスク管理・ALM",
    "02-05 事業費の管理・分析.md": "保険2第5章 事業費の管理・分析",
    "02-06 ソルベンシー.md": "保険2第6章 ソルベンシー",
    "02-07 内部管理会計.md": "保険2第7章 内部管理会計・区分経理",
    "02-08 相互会社と株式会社.md": "保険2第8章 相互会社と株式会社",
}

FULL_REF_RE = re.compile(
    r"(?P<year>H\d{1,2}|20\d{2})\s*生保[１２2]\s*(?:問(?:題)?)\s*(?P<q>\d+)(?P<subs>(?:\([^)]+\))*)"
)
SHORT_REF_RE = re.compile(r"(?P<year>H\d{1,2}|20\d{2})\s*[-－]\s*(?P<q>\d+)(?:\s*[-－]\s*(?P<sub>\d+))?")
WB_REF_LINE_RE = re.compile(r"(H\d{1,2}|20\d{2})\s*生保２問題")
CARD_HEADER_RE = re.compile(r"^## CARD\s+(.+?)\s*$", re.MULTILINE)
CARD_SOURCE_RE = re.compile(r"^- source:\s*(.+?)\s*$", re.MULTILINE)
QUESTION_SECTION_RE = re.compile(r"^### 問題\s*$", re.MULTILINE)
CHAPTER_RE = re.compile(r"保険２第([１３５６７８])章\s*(.+)")
FULLWIDTH_DIGITS = str.maketrans("１２３４５６７８９０", "1234567890")

GENERIC_TITLE_WORDS = [
    "生命保険会社",
    "生命保険",
    "日本の",
    "金融庁提出用の",
    "保険業法施行規則",
    "保険業法",
    "について",
    "における",
    "に係る",
    "場合の",
    "簡潔に",
    "説明しなさい",
    "説明せよ",
    "説明",
    "列挙しなさい",
    "列挙",
    "穴埋め問題",
    "穴埋め",
    "問題",
    "オリジナル",
    "答えは教科書",
    "記載は教科書",
    "教科書",
    "解答最新化",
    "解答アプデ",
    "長いよ",
    "とても長い",
    "まるあんき",
    "事項",
    "重要と思われる順に",
    "重要と思われる順",
    "1つ",
    "一つ",
    "筆記",
]
EXCLUDED_TITLE_PATTERNS = [
    "現行規制対象外",
    "該当する。",
    "部に計上した金額",
]


@dataclass
class WbRow:
    chapter: str
    page: int
    source_line: str
    source_refs: str
    title_hint: str


@dataclass
class NoteRow:
    chapter: str
    file: str
    title: str
    source_refs: str
    note_type: str


@dataclass
class StudyCard:
    card_id: str
    question: str
    question_norm: str
    body_norm: str
    source_refs: str
    scope: str


def normalize_ref(ref: str) -> str:
    ref = ref.replace(" ", "").replace("　", "")
    ref = ref.replace("生保2", "生保２")
    ref = ref.replace("生保1", "生保１")
    return ref


def normalize_chapter(text: str) -> str:
    return text.translate(FULLWIDTH_DIGITS).replace("保険２", "保険2")


def normalize_title(text: str) -> str:
    text = normalize_chapter(text)
    text = FULL_REF_RE.sub("", text)
    text = SHORT_REF_RE.sub("", text)
    text = re.sub(r"\b(?:H\d{1,2}|20\d{2})\b", "", text)
    text = text.replace("（", "").replace("）", "").replace("(", "").replace(")", "")
    for word in GENERIC_TITLE_WORDS:
        text = text.replace(word, "")
    text = re.sub(r"[「」『』【】［］\[\]、。,.・:：;；\s\-－_/\\]", "", text)
    return text


def extract_full_refs(text: str) -> list[str]:
    refs: list[str] = []
    for match in FULL_REF_RE.finditer(text):
        year = normalize_ref(match.group("year"))
        q = match.group("q")
        subs = re.findall(r"\(([^)]+)\)", match.group("subs"))
        numeric_subs = [sub for sub in subs if sub.isdigit()]
        if numeric_subs:
            for sub in numeric_subs:
                refs.append(f"{year}-{q}-{sub}")
        else:
            refs.append(f"{year}-{q}")
    return refs


def extract_short_refs(text: str) -> list[str]:
    refs: list[str] = []
    for match in SHORT_REF_RE.finditer(text):
        year = normalize_ref(match.group("year"))
        q = match.group("q")
        sub = match.group("sub")
        refs.append(f"{year}-{q}-{sub}" if sub else f"{year}-{q}")
    return refs


def split_refs(text: str) -> list[str]:
    refs = extract_full_refs(text) + extract_short_refs(text)
    deduped: list[str] = []
    seen: set[str] = set()
    for ref in refs:
        if ref not in seen:
            seen.add(ref)
            deduped.append(ref)
    return deduped


def refs_match(left: str, right: str) -> bool:
    left = normalize_ref(left)
    right = normalize_ref(right)
    return left == right or left.startswith(right + "-") or right.startswith(left + "-")


def any_ref_match(left_refs: list[str], right_refs: list[str]) -> bool:
    return any(refs_match(left, right) for left in left_refs for right in right_refs)


def extract_pdf_text(pdf_path: Path) -> str:
    if not PDFTOTEXT:
        raise RuntimeError(
            "pdftotext not found. Install it and add it to PATH, or set the PDFTOTEXT environment variable."
        )
    result = subprocess.run(
        [PDFTOTEXT, "-layout", "-enc", "UTF-8", str(pdf_path), "-"],
        capture_output=True,
        text=True,
        check=False,
        encoding="utf-8",
        errors="ignore",
    )
    if result.returncode != 0 and not result.stdout:
        raise RuntimeError(f"pdftotext failed for {pdf_path}")
    return result.stdout


def parse_wb() -> list[WbRow]:
    pdf = next(p for p in NOTE_DIR.glob("*.pdf") if "WB" in p.name)
    text = extract_pdf_text(pdf)
    rows: list[WbRow] = []
    current_chapter = ""
    current_page = 1
    prev_meaningful = ""

    for raw_line in text.splitlines():
        if "\f" in raw_line:
            current_page += raw_line.count("\f")
            raw_line = raw_line.replace("\f", "")
        line = raw_line.strip()
        if not line:
            continue

        chapter_match = CHAPTER_RE.search(line)
        if chapter_match:
            current_chapter = normalize_chapter(f"保険2第{chapter_match.group(1)}章 {chapter_match.group(2).strip()}")
            prev_meaningful = line
            continue

        if "生命保険会社の保険計理人の職務" in line:
            current_chapter = "生命保険会社の保険計理人の職務"
            prev_meaningful = line
            continue

        if WB_REF_LINE_RE.search(line):
            refs = split_refs(line)
            if refs:
                title_hint = prev_meaningful if prev_meaningful and "保険２第" not in prev_meaningful else ""
                rows.append(
                    WbRow(
                        chapter=current_chapter,
                        page=current_page,
                        source_line=line,
                        source_refs="; ".join(refs),
                        title_hint=title_hint,
                    )
                )
            prev_meaningful = line
            continue

        if not re.fullmatch(r"[0-9]+", line):
            prev_meaningful = line

    return rows


def parse_notes() -> list[NoteRow]:
    rows: list[NoteRow] = []
    note_dir = NOTE_DIR / "単元別マークダウン"
    for path in sorted(note_dir.glob("*.md")):
        chapter = CHAPTER_MAP.get(path.name, path.name)
        for line in path.read_text(encoding="utf-8").splitlines():
            if not line.startswith("- "):
                continue
            title = line[2:].strip()
            refs = split_refs(title)
            note_type = "original" if not refs else "past_exam"
            rows.append(
                NoteRow(
                    chapter=chapter,
                    file=path.name,
                    title=title,
                    source_refs="; ".join(refs),
                    note_type=note_type,
                )
            )
    return rows


def parse_study_sources() -> set[str]:
    refs: set[str] = set()
    study_first = STUDY_DIR / "first_part"
    if not study_first.exists():
        return refs
    for path in study_first.glob("*.md"):
        text = path.read_text(encoding="utf-8")
        for match in CARD_SOURCE_RE.finditer(text):
            refs.update(split_refs(match.group(1)))
    return refs


def extract_card_question(body: str) -> str:
    match = re.search(r"^### 問題\s*$\n(?P<question>.*?)(?=^### |\Z)", body, re.MULTILINE | re.DOTALL)
    if not match:
        return ""
    question = match.group("question").strip().splitlines()
    return question[0].strip() if question else ""


def parse_study_cards() -> list[StudyCard]:
    cards: list[StudyCard] = []
    study_first = STUDY_DIR / "first_part"
    if not study_first.exists():
        return cards
    for path in sorted(study_first.glob("*.md")):
        scope = path.name.replace("_", " ")
        text = path.read_text(encoding="utf-8")
        matches = list(CARD_HEADER_RE.finditer(text))
        for idx, match in enumerate(matches):
            body_start = match.end()
            body_end = matches[idx + 1].start() if idx + 1 < len(matches) else len(text)
            body = text[body_start:body_end]
            source_match = CARD_SOURCE_RE.search(body)
            source_text = source_match.group(1).strip() if source_match else ""
            question = extract_card_question(body)
            searchable = "\n".join([question, body])
            cards.append(
                StudyCard(
                    card_id=match.group(1).strip(),
                    question=question,
                    question_norm=normalize_title(question),
                    body_norm=normalize_title(searchable),
                    source_refs="; ".join(split_refs(source_text)),
                    scope=scope,
                )
            )
    return cards


def is_excluded_wb_row(row: WbRow) -> bool:
    title = row.title_hint or row.source_line
    if any(pattern in title for pattern in EXCLUDED_TITLE_PATTERNS):
        return True
    if len(split_refs(title)) >= 2:
        return True
    title_norm = normalize_title(title)
    if len(title_norm) <= 4 and "第" not in title:
        return True
    if title.startswith("(オ)") or title.startswith("高」「"):
        return True
    return False


def title_matches_card(title: str, card: StudyCard) -> bool:
    title_norm = normalize_title(title)
    if not title_norm:
        return False
    if title_norm == card.question_norm:
        return True
    if title_norm in card.question_norm or card.question_norm in title_norm:
        return True
    return False


def pick_best_title_match(title: str, candidates: list[StudyCard]) -> StudyCard | None:
    title_norm = normalize_title(title)
    if not title_norm:
        return None

    exact = [card for card in candidates if card.question_norm == title_norm]
    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1:
        return None

    partial = [card for card in candidates if title_matches_card(title, card)]
    if not partial:
        return None
    if len(partial) == 1:
        return partial[0]

    partial.sort(key=lambda card: len(card.question_norm), reverse=True)
    if len(partial[0].question_norm) > len(partial[1].question_norm):
        return partial[0]
    return None


def find_study_card_match(title: str, refs: list[str], study_cards: list[StudyCard], scope: str = "") -> tuple[str, str]:
    scoped_cards = [card for card in study_cards if not scope or card.scope == scope]
    if refs:
        ref_candidates = [
            card
            for card in scoped_cards
            if split_refs(card.source_refs) and any_ref_match(refs, split_refs(card.source_refs))
        ]
        if len(ref_candidates) == 1:
            return "drafted", f"ref:{ref_candidates[0].card_id}"
        best_match = pick_best_title_match(title, ref_candidates)
        if best_match:
            return "drafted", f"title:{best_match.card_id}"
        if ref_candidates:
            return "missing", ""
    best_match = pick_best_title_match(title, scoped_cards)
    if best_match:
        return "drafted", f"title:{best_match.card_id}"
    return "missing", ""


def build_canonical(wb_rows: list[WbRow], note_rows: list[NoteRow]) -> list[dict[str, str]]:
    canonical: list[dict[str, str]] = []
    used_wb_indexes: set[int] = set()
    study_cards = parse_study_cards()
    chapter_to_scope = {chapter: note_file for note_file, chapter in CHAPTER_MAP.items()}

    canonical_id = 1

    for note in note_rows:
        refs = split_refs(note.source_refs)
        matched_wb = [wb for wb in wb_rows if refs and any_ref_match(refs, split_refs(wb.source_refs))]
        for idx, wb in enumerate(wb_rows):
            if wb in matched_wb:
                used_wb_indexes.add(idx)

        origin = "notes_only"
        if matched_wb and refs:
            origin = "wb+notes"

        if refs:
            study_status, match_reason = find_study_card_match(note.title, refs, study_cards, note.file)
            needs_web_lookup = "no"
            answer_strategy = "note -> textbook/pdf -> similar local problem -> web -> reconstruct"
        else:
            study_status, match_reason = find_study_card_match(note.title, refs, study_cards, note.file)
            needs_web_lookup = "maybe"
            answer_strategy = "note -> textbook/pdf -> similar local problem -> reconstruct"

        canonical.append(
            {
                "canonical_id": f"C{canonical_id:04d}",
                "chapter": note.chapter,
                "title": note.title,
                "source_refs": "; ".join(refs),
                "origin": origin,
                "note_file": note.file,
                "study_card_id": match_reason.split(":", 1)[1] if match_reason else "",
                "match_reason": match_reason.split(":", 1)[0] if match_reason else "",
                "study_status": study_status,
                "needs_web_lookup": needs_web_lookup,
                "answer_strategy": answer_strategy,
            }
        )
        canonical_id += 1

    for idx, wb in enumerate(wb_rows):
        if idx in used_wb_indexes:
            continue
        refs = split_refs(wb.source_refs)
        if is_excluded_wb_row(wb):
            study_status = "excluded"
            match_reason = "excluded"
        else:
            scope = chapter_to_scope.get(wb.chapter, "")
            study_status, match_reason = find_study_card_match(wb.title_hint or wb.source_line, refs, study_cards, scope)
        canonical.append(
            {
                "canonical_id": f"C{canonical_id:04d}",
                "chapter": wb.chapter or "未分類",
                "title": wb.title_hint or wb.source_line,
                "source_refs": "; ".join(refs),
                "origin": "wb_only",
                "note_file": "",
                "study_card_id": match_reason.split(":", 1)[1] if ":" in match_reason else "",
                "match_reason": match_reason.split(":", 1)[0] if ":" in match_reason else match_reason,
                "study_status": study_status,
                "needs_web_lookup": "no" if study_status == "excluded" else "maybe",
                "answer_strategy": "textbook/pdf -> similar local problem -> web -> reconstruct",
            }
        )
        canonical_id += 1

    return canonical


def write_csv(rows: list[dict[str, str]], path: Path) -> None:
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def add_sheet(wb: Workbook, name: str, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet(name)
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)


def build_summary(canonical: list[dict[str, str]]) -> list[dict[str, str]]:
    chapter_counts: dict[str, dict[str, int]] = defaultdict(lambda: {"total": 0, "drafted": 0, "missing": 0, "excluded": 0})
    for row in canonical:
        chapter = row["chapter"]
        chapter_counts[chapter]["total"] += 1
        chapter_counts[chapter][row["study_status"]] += 1

    summary: list[dict[str, str]] = []
    for chapter, counts in chapter_counts.items():
        summary.append(
            {
                "chapter": chapter,
                "total": str(counts["total"]),
                "drafted": str(counts["drafted"]),
                "missing": str(counts["missing"]),
                "excluded": str(counts["excluded"]),
                "draft_rate": f"{(counts['drafted'] / counts['total'] * 100):.1f}%" if counts["total"] else "0.0%",
            }
        )
    summary.sort(key=lambda r: r["chapter"])
    return summary


def main() -> None:
    COVERAGE_DIR.mkdir(parents=True, exist_ok=True)
    wb_rows = [asdict(row) for row in parse_wb()]
    note_rows = [asdict(row) for row in parse_notes()]
    canonical = build_canonical([WbRow(**row) for row in wb_rows], [NoteRow(**row) for row in note_rows])
    summary = build_summary(canonical)

    write_csv(canonical, COVERAGE_DIR / "canonical_problems.csv")

    workbook = Workbook()
    workbook.remove(workbook.active)
    add_sheet(workbook, "summary", summary)
    add_sheet(workbook, "wb_raw", wb_rows)
    add_sheet(workbook, "notes_raw", note_rows)
    add_sheet(workbook, "canonical", canonical)
    workbook.save(COVERAGE_DIR / "problem_inventory.xlsx")

    print(f"wb_raw={len(wb_rows)}")
    print(f"notes_raw={len(note_rows)}")
    print(f"canonical={len(canonical)}")
    print(f"saved={COVERAGE_DIR / 'problem_inventory.xlsx'}")


if __name__ == "__main__":
    main()
